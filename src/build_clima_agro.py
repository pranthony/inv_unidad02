from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─────────────────────────────────────────────
# HELPER STYLES
# ─────────────────────────────────────────────
def side(style="thin", color="000000"):
    return Side(border_style=style, color=color)

def border(left=None, right=None, top=None, bottom=None):
    return Border(left=left or side(), right=right or side(),
                  top=top or side(), bottom=bottom or side())

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=9, color="000000", italic=False, name="Arial"):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# Color palette
C_TITLE_BG   = "1F4E79"   # dark blue
C_TITLE_FG   = "FFFFFF"
C_SEC_BG     = "2E75B6"   # medium blue
C_SEC_FG     = "FFFFFF"
C_HDR_BG     = "BDD7EE"   # light blue header
C_HDR_FG     = "1F4E79"
C_Y_BG       = "E2EFDA"   # green tint – dependent var
C_X_BG       = "DDEEFF"   # blue tint – independent vars
C_CALC_BG    = "F2F2F2"   # grey – calculated
C_ALT_BG     = "FAFAFA"   # very light row alt
C_NEG_BG     = "FCE4D6"   # salmon for negative expected signs
C_POS_BG     = "E2EFDA"   # green for positive expected signs
C_NOTE_BG    = "FFF2CC"   # yellow notes

def apply_cell(ws, row, col, value="", bold=False, size=9,
               fg="000000", bg=None, h="left", v="center",
               wrap=False, border_=True, italic=False, num_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font(bold=bold, size=size, color=fg, italic=italic)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if bg:
        c.fill = fill(bg)
    if border_:
        c.border = Border(left=side(), right=side(), top=side(), bottom=side())
    if num_format:
        c.number_format = num_format
    return c

def merge_title(ws, row, col_s, col_e, value, bg, fg="FFFFFF",
                size=11, bold=True):
    ws.merge_cells(start_row=row, start_column=col_s,
                   end_row=row, end_column=col_e)
    c = ws.cell(row=row, column=col_s, value=value)
    c.font = Font(bold=bold, size=size, color=fg, name="Arial")
    c.fill = fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True)
    c.border = Border(left=side("medium"), right=side("medium"),
                      top=side("medium"), bottom=side("medium"))

# ═══════════════════════════════════════════════
# SHEET 1 — FICHA TÉCNICA
# ═══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "FICHA TÉCNICA"

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 42
ws1.column_dimensions["D"].width = 28
ws1.column_dimensions["E"].width = 30
ws1.column_dimensions["F"].width = 50
ws1.column_dimensions["G"].width = 24
ws1.column_dimensions["H"].width = 30
ws1.row_dimensions[1].height = 38
ws1.row_dimensions[2].height = 18
ws1.row_dimensions[3].height = 28

# ── Row 1: Main title ─────────────────────────
merge_title(ws1, 1, 1, 8,
            "BASE DE DATOS — TEMA 1: VARIABILIDAD CLIMÁTICA Y PRODUCTIVIDAD AGRÍCOLA EN HUÁNUCO",
            C_TITLE_BG, size=13)

# ── Row 2: Subtitle ──────────────────────────
merge_title(ws1, 2, 1, 8,
            "Modelo de Series de Tiempo / MCO con Rezagos — Serie Anual 2000–2023  (24 observaciones)",
            C_SEC_BG, size=10)

# ── Row 3: Equation ──────────────────────────
merge_title(ws1, 3, 1, 8,
            "ECUACIÓN DEL MODELO ECONOMÉTRICO\n"
            "REND_PAPA_t  =  β₀  +  β₁·TEMP_t  +  β₂·PREC_t  +  β₃·PREC²_t  +  β₄·REND_PAPA_(t-1)  +  ε_t",
            "4472C4", size=10)

# ── Row 4: Blank separator ───────────────────
ws1.row_dimensions[4].height = 6

# ── Row 5: Section header FICHA ──────────────
merge_title(ws1, 5, 1, 8, "FICHA DE VARIABLES", C_SEC_BG, size=10)
ws1.row_dimensions[5].height = 22

# ── Row 6: Column headers ────────────────────
ws1.row_dimensions[6].height = 36
headers = ["ROL", "CÓDIGO", "DESCRIPCIÓN COMPLETA", "UNIDAD DE MEDIDA",
           "FUENTE", "CÓDIGO / URL", "FRECUENCIA ORIGINAL", "TRANSFORMACIÓN"]
for j, h in enumerate(headers, 1):
    apply_cell(ws1, 6, j, h, bold=True, size=9, fg=C_HDR_FG,
               bg=C_HDR_BG, h="center", wrap=True)

# ── Variable rows ────────────────────────────
vars_data = [
    # ROL, CÓDIGO, DESCRIPCIÓN, UNIDAD, FUENTE, URL, FREQ, TRANSF
    ("Y", "REND_PAPA",
     "Rendimiento del cultivo de papa en Huánuco (t/ha)",
     "Toneladas por hectárea (t/ha)",
     "MIDAGRI — Anuario Estadístico Agropecuario",
     "midagri.gob.pe → Estadística → Anuario Agropecuario → Huánuco → Cultivos Transitorios → Papa",
     "Anual", "Ninguna (usar directamente)"),

    ("X1", "TEMP_PROM",
     "Temperatura media anual en Huánuco — estaciones SENAMHI departamentales (°C)",
     "Grados Celsius (°C)",
     "SENAMHI",
     "senamhi.gob.pe → Datos Históricos → Temperatura → Huánuco → seleccionar estaciones clave (Huánuco, Llata, Tantamayo)",
     "Diaria → agregar a anual",
     "Promedio anual de datos diarios de las principales estaciones"),

    ("X2", "PREC_ANUAL",
     "Precipitación total anual acumulada en zonas de cultivo de papa en Huánuco (mm)",
     "Milímetros (mm)",
     "SENAMHI",
     "senamhi.gob.pe → Datos Históricos → Precipitación → Huánuco → estaciones en zonas andinas > 2000 m.s.n.m.",
     "Diaria / Mensual → agregar a anual",
     "Suma de precipitación mensual de los 12 meses del año"),

    ("X3", "PREC2",
     "Cuadrado de la precipitación anual — captura efecto no lineal (exceso y déficit hídrico)",
     "mm² (cuadrado de mm)",
     "Derivado de X2",
     "Calculado: PREC_ANUAL² — no requiere descarga adicional",
     "Anual (derivada)",
     "Fórmula: =PREC_ANUAL^2 en la hoja BASE DE DATOS"),

    ("X4", "REND_L1",
     "Rendimiento de papa rezagado un período (t−1) — captura inercia tecnológica y aprendizaje",
     "Toneladas por hectárea (t/ha)",
     "Derivado de Y",
     "Calculado desplazando la serie Y un año: REND_PAPA del año anterior",
     "Anual (derivada)",
     "Fórmula: =Y del año (t-1) — rezago 1"),

    ("AUX1", "MAIZ_REND",
     "Rendimiento de maíz amiláceo en Huánuco (cultivo alternativo para análisis comparado)",
     "Toneladas por hectárea (t/ha)",
     "MIDAGRI — Anuario Estadístico Agropecuario",
     "midagri.gob.pe → Anuario → Cultivos Transitorios → Maíz Amiláceo → Huánuco",
     "Anual", "Ninguna (variable auxiliar)"),

    ("AUX2", "CAFE_REND",
     "Rendimiento de café en Huánuco — zonas de selva alta (Leoncio Prado, Huamalíes)",
     "Kilogramos por hectárea (kg/ha)",
     "MIDAGRI — Anuario Estadístico Agropecuario",
     "midagri.gob.pe → Anuario → Cultivos Permanentes → Café → Huánuco",
     "Anual", "Ninguna (variable auxiliar)"),
]

row_colors = [C_Y_BG, C_X_BG, C_X_BG, C_CALC_BG, C_CALC_BG, C_ALT_BG, C_ALT_BG]

for i, (vdata, rc) in enumerate(zip(vars_data, row_colors), 7):
    ws1.row_dimensions[i].height = 42
    for j, val in enumerate(vdata, 1):
        apply_cell(ws1, i, j, val, bg=rc, h="center" if j <= 2 else "left",
                   v="center", wrap=True, size=9,
                   bold=(j == 1 or j == 2))

# ── Row 14: blank ────────────────────────────
ws1.row_dimensions[14].height = 6

# ── Signos esperados ─────────────────────────
merge_title(ws1, 15, 1, 8, "SIGNOS ESPERADOS DE LOS COEFICIENTES", C_SEC_BG, size=10)
ws1.row_dimensions[15].height = 22

sign_hdrs = ["COEFICIENTE", "SIGNO ESPERADO", "JUSTIFICACIÓN ECONÓMICA"]
ws1.merge_cells("A16:B16")
apply_cell(ws1, 16, 1, sign_hdrs[0], bold=True, bg=C_HDR_BG, h="center", fg=C_HDR_FG)
apply_cell(ws1, 16, 3, sign_hdrs[1], bold=True, bg=C_HDR_BG, h="center", fg=C_HDR_FG)
ws1.merge_cells("D16:H16")
apply_cell(ws1, 16, 4, sign_hdrs[2], bold=True, bg=C_HDR_BG, h="center", fg=C_HDR_FG)

signs = [
    ("β₁ (TEMP_PROM)", "Ambiguo (∩)",
     "Relación no lineal: temperatura moderada aumenta rendimiento; temperaturas extremas (heladas < 4°C o calor > 28°C) lo reducen. Se espera signo negativo si predominan extremos."),
    ("β₂ (PREC_ANUAL)", "Positivo (+)",
     "Mayor precipitación favorece el crecimiento de tubérculos hasta cierto umbral óptimo (~700-900 mm anuales en zona andina). Coeficiente positivo refleja efecto benéfico del agua."),
    ("β₃ (PREC²)", "Negativo (–)",
     "El término cuadrático captura el exceso hídrico: precipitaciones muy altas generan encharcamiento, erosión y enfermedades fúngicas (rancha), reduciendo el rendimiento."),
    ("β₄ (REND_L1)", "Positivo (+)",
     "La inercia tecnológica y el aprendizaje del agricultor implican que los buenos rendimientos pasados se reproducen vía mejores prácticas, semillas seleccionadas y mayor inversión."),
]

sign_colors = ["FFE5CC", C_POS_BG, C_NEG_BG, C_POS_BG]

for i, ((coef, sign, just), sc) in enumerate(zip(signs, sign_colors), 17):
    ws1.row_dimensions[i].height = 36
    ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=2)
    apply_cell(ws1, i, 1, coef, bold=True, bg=sc, h="center", v="center", wrap=True)
    apply_cell(ws1, i, 3, sign, bold=True, bg=sc, h="center", v="center")
    ws1.merge_cells(start_row=i, start_column=4, end_row=i, end_column=8)
    apply_cell(ws1, i, 4, just, bg=C_ALT_BG, h="left", v="center", wrap=True)

# ── Notas metodológicas ──────────────────────
ws1.row_dimensions[22].height = 6
merge_title(ws1, 23, 1, 8, "NOTAS METODOLÓGICAS", C_SEC_BG, size=10)
ws1.row_dimensions[23].height = 22

notas = [
    "1. Período de análisis: 2000–2023 (24 años). Mínimo recomendado para MCO con 4 regresores y rezago.",
    "2. La variable Y proviene de los Anuarios Estadísticos Agropecuarios de MIDAGRI (antes MINAGRI). Se prioriza la serie de papa por ser el principal cultivo andino de Huánuco (CENAGRO 2012).",
    "3. Las variables climáticas (X1 y X2) provienen de SENAMHI. Usar el promedio de 2-3 estaciones representativas de la zona papera del departamento (Huánuco, Llata, Tantamayo). Verificar con CENAGRO la distribución espacial.",
    "4. X3 (PREC²) es generada internamente en Excel con fórmula automática. No requiere fuente externa.",
    "5. X4 (rezago) desplaza la serie Y un período. La primera observación (2000) perderá el rezago; el modelo efectivo tendrá 23 observaciones (2001–2023) si se incluye el rezago.",
    "6. Verificar estacionariedad con prueba ADF/PP antes de estimar. Series climáticas pueden tener tendencia por cambio climático. Si hay raíz unitaria, trabajar con primeras diferencias.",
    "7. Estimar con MCO. Evaluar autocorrelación (Durbin-Watson), heteroscedasticidad (Breusch-Pagan) y multicolinealidad (VIF < 10).",
    "8. Para robustez: estimar el modelo también para maíz (MAIZ_REND) y café (CAFE_REND) como variables auxiliares y comparar coeficientes climáticos entre cultivos.",
    "9. Software recomendado: EViews, Stata, R (paquete dynlm + lmtest) o Python (statsmodels). En Excel puede usarse Analysis ToolPak para MCO básico.",
]

for i, nota in enumerate(notas, 24):
    ws1.row_dimensions[i + 1].height = 30
    ws1.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
    c = ws1.cell(row=i, column=1, value=nota)
    c.font = Font(size=9, name="Arial", italic=(i == 24))
    c.fill = fill(C_NOTE_BG if i % 2 == 0 else "FFFDE7")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = Border(left=side(), right=side(), top=side(), bottom=side())
    ws1.row_dimensions[i].height = 30

# ═══════════════════════════════════════════════
# SHEET 2 — BASE DE DATOS
# ═══════════════════════════════════════════════
ws2 = wb.create_sheet("BASE DE DATOS")

col_widths = [14, 7, 17, 18, 18, 18, 18, 18, 18, 18]
col_letters = [get_column_letter(i+1) for i in range(len(col_widths))]
for cl, cw in zip(col_letters, col_widths):
    ws2.column_dimensions[cl].width = cw

ws2.row_dimensions[1].height = 38
ws2.row_dimensions[2].height = 18
ws2.row_dimensions[3].height = 50
ws2.row_dimensions[4].height = 50

# ── Titles ────────────────────────────────────
merge_title(ws2, 1, 1, 10,
            "BASE DE DATOS — TEMA 1: VARIABILIDAD CLIMÁTICA Y PRODUCTIVIDAD AGRÍCOLA EN HUÁNUCO",
            C_TITLE_BG, size=12)

merge_title(ws2, 2, 1, 10,
            "Serie anual 2000–2023 | Fuentes: SENAMHI, MIDAGRI-Anuario Agropecuario, CENAGRO | Completar celdas en verde (Y) y azul (X)",
            C_SEC_BG, size=9)

# ── Group headers row 3 ───────────────────────
group_headers = [
    (1, 2, "IDENTIFICACIÓN", C_HDR_BG),
    (3, 3, "VARIABLE DEPENDIENTE", "C6EFCE"),
    (4, 7, "VARIABLES INDEPENDIENTES", "BDD7EE"),
    (8, 10, "VARIABLES AUXILIARES / CALCULADAS", C_CALC_BG),
]
for cs, ce, label, bg_c in group_headers:
    ws2.merge_cells(start_row=3, start_column=cs, end_row=3, end_column=ce)
    c = ws2.cell(row=3, column=cs, value=label)
    c.font = Font(bold=True, size=9, color=C_HDR_FG, name="Arial")
    c.fill = fill(bg_c)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(left=side("medium"), right=side("medium"),
                      top=side("medium"), bottom=side("medium"))

# ── Column headers row 4 ─────────────────────
col_headers = [
    ("AÑO\n(Período)", "A4", None),
    ("AÑO\n(Num.)", "B4", None),
    ("Y\nREND_PAPA\n(t/ha)", "C4", "C6EFCE"),
    ("X1\nTEMP_PROM\n(°C)", "D4", "BDD7EE"),
    ("X2\nPREC_ANUAL\n(mm)", "E4", "BDD7EE"),
    ("X3\nPREC²\n(mm²)", "F4", C_CALC_BG),
    ("X4\nREND_L1\n(t/ha t-1)", "G4", C_CALC_BG),
    ("AUX1\nMAIZ_REND\n(t/ha)", "H4", "F2F2F2"),
    ("AUX2\nCAFE_REND\n(kg/ha)", "I4", "F2F2F2"),
    ("NOTAS\n/ANOMALÍAS", "J4", "FFF2CC"),
]
for j, (label, _, bg_c) in enumerate(col_headers, 1):
    c = ws2.cell(row=4, column=j, value=label)
    c.font = Font(bold=True, size=8, color=C_HDR_FG, name="Arial")
    c.fill = fill(bg_c or C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = Border(left=side("medium"), right=side("medium"),
                      top=side("medium"), bottom=side("medium"))

# ── Data rows 2000-2023 ───────────────────────
years = list(range(2000, 2024))
anomalies = {
    2004: "Año Niño moderado",
    2010: "Heladas severas zona andina",
    2017: "Lluvias extremas — Niño Costero",
    2020: "COVID-19: limitaciones de datos",
    2022: "Inundaciones andinas",
}

for i, yr in enumerate(years, 5):
    row = i
    ws2.row_dimensions[row].height = 18
    alt = "FFFFFF" if i % 2 == 0 else "F5FBFF"

    # A: period label
    c = ws2.cell(row=row, column=1, value=str(yr))
    c.font = Font(bold=True, size=9, name="Arial", color="1F4E79")
    c.fill = fill("EBF3FB")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=side("medium"), right=side(), top=side(), bottom=side())

    # B: year number
    c = ws2.cell(row=row, column=2, value=yr)
    c.font = Font(size=9, name="Arial")
    c.fill = fill("EBF3FB")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=side(), right=side(), top=side(), bottom=side())
    c.number_format = "0"

    # C: Y (manual – green)
    c = ws2.cell(row=row, column=3, value=None)
    c.fill = fill("E2EFDA")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=side("medium"), right=side(), top=side(), bottom=side())
    c.number_format = "0.00"

    # D: X1 (manual – blue)
    c = ws2.cell(row=row, column=4, value=None)
    c.fill = fill("DDEEFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=side(), right=side(), top=side(), bottom=side())
    c.number_format = "0.0"

    # E: X2 (manual – blue)
    c = ws2.cell(row=row, column=5, value=None)
    c.fill = fill("DDEEFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=side(), right=side(), top=side(), bottom=side())
    c.number_format = "0.0"

    # F: X3 = PREC² (formula – grey)
    e_col = f"E{row}"
    formula_prec2 = f'=IF(ISNUMBER({e_col}),{e_col}^2,"")'
    c = ws2.cell(row=row, column=6, value=formula_prec2)
    c.fill = fill(C_CALC_BG)
    c.font = Font(size=9, name="Arial", color="595959")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=side(), right=side(), top=side(), bottom=side())
    c.number_format = "#,##0"

    # G: X4 = REND_L1 – lag of Y (formula – grey; first year is empty)
    if i == 5:
        c = ws2.cell(row=row, column=7, value="N/A (sin rezago)")
        c.fill = fill(C_CALC_BG)
        c.font = Font(size=8, name="Arial", color="595959", italic=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=side(), right=side(), top=side(), bottom=side())
    else:
        prev_c_row = row - 1
        formula_lag = f'=IF(ISNUMBER(C{prev_c_row}),C{prev_c_row},"")'
        c = ws2.cell(row=row, column=7, value=formula_lag)
        c.fill = fill(C_CALC_BG)
        c.font = Font(size=9, name="Arial", color="595959")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=side(), right=side(), top=side(), bottom=side())
        c.number_format = "0.00"

    # H: AUX1 MAIZ (manual – light grey)
    c = ws2.cell(row=row, column=8, value=None)
    c.fill = fill("F5F5F5")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=side("medium"), right=side(), top=side(), bottom=side())
    c.number_format = "0.00"

    # I: AUX2 CAFE (manual – light grey)
    c = ws2.cell(row=row, column=9, value=None)
    c.fill = fill("F5F5F5")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(left=side(), right=side(), top=side(), bottom=side())
    c.number_format = "0"

    # J: Notes
    note_val = anomalies.get(yr, "")
    c = ws2.cell(row=row, column=10, value=note_val)
    c.fill = fill("FFF9E6" if note_val else alt)
    c.font = Font(size=8, name="Arial", italic=bool(note_val),
                  color="7F6000" if note_val else "000000")
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c.border = Border(left=side(), right=side("medium"), top=side(), bottom=side())

# ── Summary statistics ────────────────────────
stat_start = 5 + len(years) + 1  # row 30

stat_labels = [
    ("N (obs.)",  None,         None),
    ("Media",     "=AVERAGE(",  "0.00"),
    ("Mín.",      "=MIN(",      "0.00"),
    ("Máx.",      "=MAX(",      "0.00"),
    ("D. Estándar", "=STDEV(", "0.00"),
]
stat_cols = ["C", "D", "E", "H", "I"]

for k, (label, func, nfmt) in enumerate(stat_labels, stat_start):
    ws2.row_dimensions[k].height = 20
    ws2.merge_cells(start_row=k, start_column=1, end_row=k, end_column=2)
    c = ws2.cell(row=k, column=1, value=label)
    c.font = Font(bold=True, size=9, name="Arial", color="1F4E79")
    c.fill = fill("D6E4F0")
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = Border(left=side("medium"), right=side(), top=side("medium"),
                      bottom=side())

    for sc in range(3, 11):
        col_letter = get_column_letter(sc)
        if func is None:
            val = f'=COUNTA({col_letter}5:{col_letter}{stat_start-2})' if sc < 8 else ""
        elif col_letter in stat_cols:
            val = f'{func}{col_letter}5:{col_letter}{stat_start-2})'
        else:
            val = ""
        c2 = ws2.cell(row=k, column=sc, value=val if val else None)
        c2.fill = fill("D6E4F0")
        c2.font = Font(size=9, name="Arial", color="000000")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.border = Border(left=side(), right=side("medium") if sc == 10 else side(),
                           top=side("medium"), bottom=side())
        if nfmt:
            c2.number_format = nfmt

# ── Legend ────────────────────────────────────
leg_row = stat_start + len(stat_labels) + 1
ws2.merge_cells(start_row=leg_row, start_column=1, end_row=leg_row, end_column=10)
c = ws2.cell(row=leg_row, column=1,
             value="LEYENDA DE COLORES:  🟢 Verde = Variable dependiente Y (ingresar manualmente)   "
                   "🔵 Azul = Variables independientes X1, X2 (ingresar manualmente)   "
                   "⬜ Gris = Fórmulas calculadas automáticamente (X3 = PREC², X4 = REND_L1)   "
                   "🟡 Amarillo = Nota de anomalía climática o metodológica")
c.font = Font(size=8, name="Arial", italic=True, color="595959")
c.fill = fill("FFFDE7")
c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
c.border = Border(left=side("medium"), right=side("medium"),
                  top=side("medium"), bottom=side("medium"))
ws2.row_dimensions[leg_row].height = 30

# freeze panes
ws2.freeze_panes = "C5"

# ═══════════════════════════════════════════════
# SHEET 3 — GUÍA DE DESCARGA
# ═══════════════════════════════════════════════
ws3 = wb.create_sheet("GUÍA DE DESCARGA")
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 28
ws3.column_dimensions["C"].width = 72
ws3.row_dimensions[1].height = 36

merge_title(ws3, 1, 1, 3, "GUÍA PASO A PASO PARA DESCARGA DE DATOS — TEMA 1: VARIABILIDAD CLIMÁTICA Y PRODUCTIVIDAD AGRÍCOLA", C_TITLE_BG, size=11)

guides = [
    ("Y — REND_PAPA | Fuente: MIDAGRI — Anuario Estadístico Agropecuario", "2E75B6", [
        ("1", "Ir a la web de MIDAGRI", "https://www.midagri.gob.pe"),
        ("2", "Ruta de navegación", "Estadística → Anuario Estadístico Agropecuario → Seleccionar el año → Descargar PDF o Excel del anuario"),
        ("3", "Filtrar por departamento y cultivo", "En el anuario busca la sección 'Cultivos Transitorios' → Papa → Huánuco. Extrae la columna 'Rendimiento (kg/ha)' o 't/ha'."),
        ("4", "Convertir unidades", "Si los datos están en kg/ha, dividir entre 1000 para obtener t/ha. Anotar el año cosecha (año t), no el año de campaña."),
        ("5", "Fuente alternativa — CENAGRO", "INEI → CENAGRO 2012 → Tabulados departamentales → Huánuco → cultivos → superficie, producción, rendimiento."),
    ]),
    ("X1 — TEMP_PROM | Fuente: SENAMHI", "2E75B6", [
        ("1", "Ir al portal SENAMHI", "https://www.senamhi.gob.pe"),
        ("2", "Ruta de navegación", "Servicios → Datos Meteorológicos Históricos → Temperatura → Departamento: Huánuco → Seleccionar estaciones"),
        ("3", "Estaciones recomendadas", "Priorizar: Huánuco (HCO-001), Llata, Tantamayo, Jesús. Estas cubren los pisos altitudinales donde se cultiva papa (2500–3800 msnm)."),
        ("4", "Periodo y frecuencia", "Descargar datos mensuales de temperatura media (T_med) para 2000–2023. El portal permite descarga en CSV o Excel por estación."),
        ("5", "Agregar a anual", "Calcular el promedio anual de T_med de los 12 meses. Si hay varias estaciones, calcular el promedio entre ellas (promedio espacial)."),
        ("6", "Datos faltantes", "Estaciones pueden tener gaps. Usar interpolación lineal para años con datos faltantes o reemplazar con datos de la estación más cercana."),
    ]),
    ("X2 — PREC_ANUAL | Fuente: SENAMHI", "2E75B6", [
        ("1", "Ir al portal SENAMHI", "https://www.senamhi.gob.pe → Datos Históricos → Precipitación → Huánuco"),
        ("2", "Seleccionar estaciones", "Usar las mismas estaciones que para temperatura. La precipitación es más variable espacialmente; usar al menos 3 estaciones."),
        ("3", "Dato a extraer", "Precipitación total mensual (PP_total en mm). Suma de precipitación diaria del mes."),
        ("4", "Agregar a anual", "Sumar los 12 meses para obtener precipitación total anual (mm/año). Anotar el año hidrológico correspondiente."),
        ("5", "Fuente alternativa", "CHIRPS (Climate Hazards Group InfraRed Precipitation with Station Data): datos grillados de alta resolución. Descargar via GIOVANNI-NASA o Google Earth Engine para coordenadas de Huánuco."),
        ("6", "Validación", "Precipitación media anual esperada en zona papera andina de Huánuco: 600–1000 mm. Valores muy fuera de este rango pueden indicar error de datos."),
    ]),
    ("X3 — PREC² | Calculado automáticamente en Excel", "4472C4", [
        ("1", "No requiere descarga", "Esta variable se genera automáticamente en la hoja BASE DE DATOS con la fórmula: =E(fila)^2, donde E corresponde a la columna PREC_ANUAL."),
        ("2", "Propósito econométrico", "El término cuadrático permite modelar la relación no lineal entre precipitación y rendimiento: rendimiento óptimo a nivel intermedio de lluvia (forma de U invertida)."),
        ("3", "Interpretación del par β₂/β₃", "Si β₂ > 0 y β₃ < 0: existe un umbral óptimo de precipitación = −β₂/(2·β₃). Calcular e interpretar agronómicamente."),
    ]),
    ("X4 — REND_L1 (Rezago) | Calculado automáticamente", "4472C4", [
        ("1", "No requiere descarga", "El rezago se genera automáticamente en la hoja BASE DE DATOS: la celda del año t toma el valor de REND_PAPA del año t-1."),
        ("2", "Importante para el modelo", "El rezago captura inercia del sistema agrícola. Eliminar la primera observación (año 2000) del modelo estimado, pues no tiene rezago disponible."),
        ("3", "Prueba de endogeneidad", "Si se sospecha endogeneidad del rezago, aplicar prueba de Hausman o usar variables instrumentales en estimación robusta."),
    ]),
]

current_row = 2
for section_title, sec_color, steps in guides:
    ws3.row_dimensions[current_row].height = 24
    ws3.merge_cells(start_row=current_row, start_column=1,
                    end_row=current_row, end_column=3)
    c = ws3.cell(row=current_row, column=1, value=section_title)
    c.font = Font(bold=True, size=10, color="FFFFFF", name="Arial")
    c.fill = fill(sec_color)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(left=side("medium"), right=side("medium"),
                      top=side("medium"), bottom=side("medium"))
    current_row += 1

    # step header
    for j, hdr in enumerate(["PASO", "ACCIÓN", "DETALLE"], 1):
        c = ws3.cell(row=current_row, column=j, value=hdr)
        c.font = Font(bold=True, size=9, color=C_HDR_FG, name="Arial")
        c.fill = fill(C_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=side(), right=side(), top=side(), bottom=side())
    ws3.row_dimensions[current_row].height = 18
    current_row += 1

    for step_num, action, detail in steps:
        ws3.row_dimensions[current_row].height = 34
        alt = "F5FBFF" if int(step_num) % 2 == 1 else "FFFFFF"
        for j, val in enumerate([step_num, action, detail], 1):
            c = ws3.cell(row=current_row, column=j, value=val)
            c.font = Font(size=9, name="Arial",
                          bold=(j == 1), color="1F4E79" if j == 1 else "000000")
            c.fill = fill(alt)
            c.alignment = Alignment(horizontal="center" if j == 1 else "left",
                                    vertical="center", wrap_text=True)
            c.border = Border(left=side("medium") if j == 1 else side(),
                              right=side("medium") if j == 3 else side(),
                              top=side(), bottom=side())
        current_row += 1

    current_row += 1  # blank row between sections

# ═══════════════════════════════════════════════
# SHEET 4 — MODELO ECONOMÉTRICO
# ═══════════════════════════════════════════════
ws4 = wb.create_sheet("MODELO ECONOMÉTRICO")
ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 78
ws4.row_dimensions[1].height = 36

merge_title(ws4, 1, 1, 2,
            "ESQUEMA DEL MODELO ECONOMÉTRICO — TEMA 1: VARIABILIDAD CLIMÁTICA Y PRODUCTIVIDAD AGRÍCOLA",
            C_TITLE_BG, size=11)

sections_model = [
    ("ESPECIFICACIÓN DEL MODELO", C_SEC_BG, [
        ("Tipo de modelo", "Regresión lineal múltiple con rezago por Mínimos Cuadrados Ordinarios (MCO)"),
        ("Forma funcional", "Cuadrático-lineal con rezago: Y en niveles; X3 = PREC² captura no linealidad; X4 = Y(t-1) como variable de inercia"),
        ("Ecuación estimable", "REND_PAPA_t = β₀ + β₁·TEMP_PROM_t + β₂·PREC_ANUAL_t + β₃·PREC²_t + β₄·REND_PAPA_(t-1) + ε_t"),
        ("Número de observaciones", "23 años efectivos (2001–2023), pues el primer año (2000) no tiene rezago disponible"),
        ("Número de parámetros", "5 (β₀, β₁, β₂, β₃, β₄)"),
        ("Grados de libertad", "18 (suficiente para inferencia básica con 5 parámetros)"),
    ]),
    ("PASOS DE ESTIMACIÓN", C_SEC_BG, [
        ("Paso 1 — Análisis descriptivo", "Calcular media, desviación estándar, mínimo y máximo de cada variable. Graficar series de rendimiento y clima en el tiempo para detectar tendencias, ciclos (ENSO) y outliers. Identificar años con anomalías climáticas extremas."),
        ("Paso 2 — Análisis de correlaciones", "Construir matriz de correlaciones entre Y, X1, X2, X3, X4. Verificar correlación entre X2 (PREC) y X3 (PREC²) — puede haber multicolinealidad. Si |r(X2,X3)| > 0.90, centrar la variable: usar (PREC − media_PREC) antes de elevar al cuadrado."),
        ("Paso 3 — Prueba de estacionariedad", "Aplicar prueba ADF o KPSS a REND_PAPA y TEMP_PROM. Las series climáticas con tendencia por cambio climático pueden no ser estacionarias. Si se rechaza estacionariedad → usar primeras diferencias o incluir tendencia temporal (t) como regresor adicional."),
        ("Paso 4 — Estimación MCO", "Estimar la ecuación por OLS. Obtener: coeficientes β̂, errores estándar robustos (Newey-West si hay autocorrelación), t-estadísticos, p-valores, R², R² ajustado y F-estadístico global. Calcular umbral óptimo de precipitación: PREC* = −β₂/(2·β₃)."),
        ("Paso 5 — Diagnóstico de residuos", "a) Autocorrelación: Breusch-Godfrey LM test (superior a DW en presencia de rezago). Si p < 0.05 → aplicar errores Newey-West.\nb) Heteroscedasticidad: Breusch-Pagan (H₀: homosc.). Si se rechaza → errores robustos HC.\nc) Normalidad: Jarque-Bera.\nd) Estabilidad: prueba CUSUM o Chow si hay sospecha de quiebre estructural (ej. 2010 o 2017)."),
        ("Paso 6 — Interpretación agronómica", "β₁: variación en t/ha ante aumento de 1°C (efecto temperatura).\nβ₂ y β₃: permiten calcular precipitación óptima PREC* = −β₂/(2·β₃) en mm/año.\nβ₄: coeficiente de inercia (persistencia inter-anual del rendimiento).\nR²: proporción de variabilidad del rendimiento explicada por el clima y el rezago."),
    ]),
    ("CRITERIOS DE BONDAD DE AJUSTE", "4472C4", [
        ("R² ajustado", "Debe ser > 0.55 para datos agroclimáticos anuales regionales (alta variabilidad natural)."),
        ("F-estadístico (global)", "p-valor < 0.05 → el modelo en su conjunto es significativo."),
        ("t-estadísticos por variable", "p-valor < 0.10 es aceptable con 23 obs.; p < 0.05 es preferible."),
        ("Prueba Breusch-Godfrey", "p-valor > 0.05 indica ausencia de autocorrelación en residuos (importante con rezago en modelo)."),
        ("VIF (Variance Inflation Factor)", "VIF < 10 para cada regresor; atención especial al par PREC y PREC² (centrar si VIF > 10)."),
        ("Umbral óptimo PREC*", "−β₂/(2·β₃) debe estar dentro del rango observado de precipitación (600–1000 mm) para tener sentido agronómico."),
    ]),
    ("HIPÓTESIS DE LA INVESTIGACIÓN", C_SEC_BG, [
        ("Hipótesis general", "La variabilidad climática (temperatura y precipitación) influye significativamente en el rendimiento del cultivo de papa en el departamento de Huánuco durante el período 2000–2023, con una relación no lineal respecto a la precipitación."),
        ("H₁ (X1: Temperatura)", "La temperatura media anual tiene efecto significativo sobre el rendimiento de papa en Huánuco. En el rango observado, se espera efecto negativo si predominan eventos de helada o calor extremo (β₁ ≠ 0)."),
        ("H₂/H₃ (X2 y X3: Precipitación + cuadrático)", "Existe un nivel óptimo de precipitación (PREC*) que maximiza el rendimiento de papa. Por encima o por debajo de este umbral, el rendimiento decrece (β₂ > 0; β₃ < 0)."),
        ("H₄ (X4: Rezago)", "El rendimiento del año anterior influye positivamente en el rendimiento actual, reflejando inercia tecnológica, calidad de semilla y acumulación de capital del agricultor (β₄ > 0)."),
    ]),
]

current_row_m = 2
for sec_title, sec_bg, rows in sections_model:
    ws4.row_dimensions[current_row_m].height = 24
    ws4.merge_cells(start_row=current_row_m, start_column=1,
                    end_row=current_row_m, end_column=2)
    c = ws4.cell(row=current_row_m, column=1, value=sec_title)
    c.font = Font(bold=True, size=10, color="FFFFFF", name="Arial")
    c.fill = fill(sec_bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = Border(left=side("medium"), right=side("medium"),
                      top=side("medium"), bottom=side("medium"))
    current_row_m += 1

    for k, (label, detail) in enumerate(rows):
        ws4.row_dimensions[current_row_m].height = 44
        alt = "EBF3FB" if k % 2 == 0 else "FFFFFF"
        c1 = ws4.cell(row=current_row_m, column=1, value=label)
        c1.font = Font(bold=True, size=9, name="Arial", color="1F4E79")
        c1.fill = fill(alt)
        c1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c1.border = Border(left=side("medium"), right=side(), top=side(), bottom=side())

        c2 = ws4.cell(row=current_row_m, column=2, value=detail)
        c2.font = Font(size=9, name="Arial")
        c2.fill = fill(alt)
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c2.border = Border(left=side(), right=side("medium"), top=side(), bottom=side())
        current_row_m += 1

    current_row_m += 1  # blank

wb.save("Tema2_ClimaAgro_Huanuco.xlsx")
print("✅ Archivo guardado correctamente.")
