"""
Cálculo del IPC (Índice de Precios al Consumidor) - Nivel Nacional
Fuente: INEI - Base Diciembre 2021 = 100
Calcula: IPC anual promedio y IPC trimestral promedio
"""

import pandas as pd
import numpy as np

FILE_PATH = "ipc_data.xlsx"
SHEET    = "Base Dic.2021"

# ──────────────────────────────────────────────
# 1. CARGA Y LIMPIEZA DE DATOS
# ──────────────────────────────────────────────
raw = pd.read_excel(FILE_PATH, sheet_name=SHEET, header=None)

# Renombrar columnas usando la fila 3 como cabecera (índice 3)
raw.columns = ["Año", "Mes", "Índice", "Var_Mensual_%", "Var_Acumulada_%", "Var_Anual_%"]
raw = raw.iloc[4:].reset_index(drop=True)   # quitar encabezados

# Forward-fill la columna Año (los NaN heredan el año anterior)
raw["Año"] = raw["Año"].ffill()
raw["Año"] = raw["Año"].astype(int)

# Convertir Índice a numérico (por si hubiera strings)
raw["Índice"] = pd.to_numeric(raw["Índice"], errors="coerce")
raw = raw.dropna(subset=["Índice", "Mes"])

# ──────────────────────────────────────────────
# 2. MAPEO DE MES A NÚMERO Y TRIMESTRE
# ──────────────────────────────────────────────
MESES = {
    "Enero": 1, "Febrero": 2, "Marzo": 3, "Abril": 4,
    "Mayo": 5, "Junio": 6, "Julio": 7, "Agosto": 8,
    "Setiembre": 9, "Septiembre": 9, "Octubre": 10,
    "Noviembre": 11, "Diciembre": 12
}

raw["Mes_Num"]   = raw["Mes"].map(MESES)
raw["Trimestre"] = ((raw["Mes_Num"] - 1) // 3 + 1).astype(int)

# ──────────────────────────────────────────────
# 3. IPC ANUAL — promedio de índices mensuales
# ──────────────────────────────────────────────
ipc_anual = (
    raw.groupby("Año")["Índice"]
    .agg(
        IPC_Promedio="mean",
        IPC_Min="min",
        IPC_Max="max",
        Meses_Disponibles="count"
    )
    .reset_index()
)

# Variación anual promedio respecto al año anterior
ipc_anual["Var_Anual_%"] = (
    (ipc_anual["IPC_Promedio"] / ipc_anual["IPC_Promedio"].shift(1) - 1) * 100
).round(4)

ipc_anual["IPC_Promedio"] = ipc_anual["IPC_Promedio"].round(4)
ipc_anual["IPC_Min"]      = ipc_anual["IPC_Min"].round(4)
ipc_anual["IPC_Max"]      = ipc_anual["IPC_Max"].round(4)

# ──────────────────────────────────────────────
# 4. IPC TRIMESTRAL — promedio por año-trimestre
# ──────────────────────────────────────────────
ipc_trim = (
    raw.groupby(["Año", "Trimestre"])["Índice"]
    .agg(
        IPC_Promedio="mean",
        IPC_Min="min",
        IPC_Max="max",
        Meses_Disponibles="count"
    )
    .reset_index()
)

ipc_trim["IPC_Promedio"] = ipc_trim["IPC_Promedio"].round(4)
ipc_trim["IPC_Min"]      = ipc_trim["IPC_Min"].round(4)
ipc_trim["IPC_Max"]      = ipc_trim["IPC_Max"].round(4)
ipc_trim["Trimestre_Nombre"] = "Q" + ipc_trim["Trimestre"].astype(str)

# Variación trimestral respecto al trimestre anterior
ipc_trim = ipc_trim.sort_values(["Año", "Trimestre"]).reset_index(drop=True)
ipc_trim["Var_Trim_%"] = (
    (ipc_trim["IPC_Promedio"] / ipc_trim["IPC_Promedio"].shift(1) - 1) * 100
).round(4)
# Limpiar variación cuando cambia de año para trimestres iniciales incompletos
# (se mantiene, pues la serie es continua)

# ──────────────────────────────────────────────
# 5. IMPRESIÓN DE RESULTADOS
# ──────────────────────────────────────────────
SEP = "=" * 80

print(SEP)
print("  ÍNDICE DE PRECIOS AL CONSUMIDOR — NIVEL NACIONAL  (Base Dic. 2021 = 100)")
print("  Fuente: INEI  |  Cálculos: Python")
print(SEP)

# ── Anual ──────────────────────────────────────
print("\n📅  IPC ANUAL  (Promedio de índices mensuales)")
print("-" * 72)
print(f"{'Año':>6}  {'IPC Prom.':>11}  {'Mín.':>10}  {'Máx.':>10}  "
      f"{'Meses':>6}  {'Var. Anual %':>13}")
print("-" * 72)
for _, r in ipc_anual.iterrows():
    var = f"{r['Var_Anual_%']:>+12.2f}%" if pd.notna(r["Var_Anual_%"]) else f"{'—':>13}"
    print(f"{int(r['Año']):>6}  {r['IPC_Promedio']:>11.4f}  "
          f"{r['IPC_Min']:>10.4f}  {r['IPC_Max']:>10.4f}  "
          f"{int(r['Meses_Disponibles']):>6}  {var}")
print("-" * 72)

# ── Trimestral ─────────────────────────────────
print("\n\n📊  IPC TRIMESTRAL  (Promedio de índices del trimestre)")
print("-" * 72)
print(f"{'Año':>6}  {'Trim.':>5}  {'IPC Prom.':>11}  {'Mín.':>10}  "
      f"{'Máx.':>10}  {'Meses':>6}  {'Var. Trim. %':>13}")
print("-" * 72)
for _, r in ipc_trim.iterrows():
    var = f"{r['Var_Trim_%']:>+12.2f}%" if pd.notna(r["Var_Trim_%"]) else f"{'—':>13}"
    print(f"{int(r['Año']):>6}  {r['Trimestre_Nombre']:>5}  "
          f"{r['IPC_Promedio']:>11.4f}  {r['IPC_Min']:>10.4f}  "
          f"{r['IPC_Max']:>10.4f}  {int(r['Meses_Disponibles']):>6}  {var}")
print("-" * 72)

# ── Resumen estadístico ────────────────────────
print("\n\n📈  RESUMEN ESTADÍSTICO DEL PERÍODO COMPLETO")
print("-" * 50)
print(f"  Período analizado : {int(ipc_anual['Año'].min())} – {int(ipc_anual['Año'].max())}")
print(f"  Observaciones     : {len(raw)} meses")
print(f"  IPC más bajo      : {raw['Índice'].min():.4f}  "
      f"({raw.loc[raw['Índice'].idxmin(), 'Mes']} {int(raw.loc[raw['Índice'].idxmin(), 'Año'])})")
print(f"  IPC más alto      : {raw['Índice'].max():.4f}  "
      f"({raw.loc[raw['Índice'].idxmax(), 'Mes']} {int(raw.loc[raw['Índice'].idxmax(), 'Año'])})")
print(f"  Var. anual máx.   : "
      f"{ipc_anual['Var_Anual_%'].max():+.2f}%  "
      f"(año {int(ipc_anual.loc[ipc_anual['Var_Anual_%'].idxmax(), 'Año'])})")
print(f"  Var. anual mín.   : "
      f"{ipc_anual['Var_Anual_%'].min():+.2f}%  "
      f"(año {int(ipc_anual.loc[ipc_anual['Var_Anual_%'].idxmin(), 'Año'])})")
print("-" * 50)

# ──────────────────────────────────────────────
# 6. EXPORTAR A EXCEL
# ──────────────────────────────────────────────
OUTPUT = "IPC_Analisis.xlsx"

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Primero guardar con pandas
with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
    ipc_anual[["Año", "IPC_Promedio", "IPC_Min", "IPC_Max",
               "Meses_Disponibles", "Var_Anual_%"]].to_excel(
        writer, sheet_name="IPC Anual", index=False)
    ipc_trim[["Año", "Trimestre_Nombre", "IPC_Promedio", "IPC_Min",
              "IPC_Max", "Meses_Disponibles", "Var_Trim_%"]].to_excel(
        writer, sheet_name="IPC Trimestral", index=False)
    raw[["Año", "Mes", "Índice", "Var_Mensual_%",
         "Var_Acumulada_%", "Var_Anual_%"]].to_excel(
        writer, sheet_name="Datos Mensuales", index=False)

# Aplicar formato
wb = load_workbook(OUTPUT)

HEADER_FILL  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
ROW_ALT_FILL = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
HEADER_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
CELL_FONT    = Font(name="Arial", size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center")
thin_side    = Side(style="thin", color="BFBFBF")
THIN_BORDER  = Border(left=thin_side, right=thin_side,
                      top=thin_side, bottom=thin_side)

def format_sheet(ws, col_widths):
    # Encabezados
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
    # Filas de datos
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ROW_ALT_FILL if i % 2 == 0 else PatternFill()
        for cell in row:
            cell.font      = CELL_FONT
            cell.fill      = fill
            cell.alignment = CENTER
            cell.border    = THIN_BORDER
    # Anchos de columna
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    # Altura de encabezado
    ws.row_dimensions[1].height = 22

format_sheet(wb["IPC Anual"],      [8, 14, 12, 12, 10, 14])
format_sheet(wb["IPC Trimestral"], [8, 8, 14, 12, 12, 10, 15])
format_sheet(wb["Datos Mensuales"],[8, 13, 12, 15, 17, 13])

wb.save(OUTPUT)
print(f"\n✅  Archivo Excel guardado en:\n   {OUTPUT}")