from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color palette ──────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"   # header bg
MID_BLUE    = "2E75B6"   # section header bg
LIGHT_BLUE  = "D6E4F0"   # Y variable row bg
LIGHT_GREEN = "E2EFDA"   # X variable rows bg
LIGHT_GRAY  = "F2F2F2"   # calculated / aux rows bg
YELLOW_HL   = "FFF2CC"   # notes / highlight
ORANGE_HDR  = "C65911"   # sub-headers (signs table)
WHITE       = "FFFFFF"
BLUE_TEXT   = "0070C0"   # blue (hard-coded inputs)
BLACK_TEXT  = "000000"
GREEN_TEXT  = "375623"
RED_TEXT    = "C00000"

def hdr(text, bold=True, sz=11, color=WHITE, bg=None, wrap=False, italic=False):
    cell_style = {"font": Font(name="Arial", bold=bold, size=sz,
                               color=color, italic=italic),
                  "alignment": Alignment(horizontal="center",
                                         vertical="center",
                                         wrap_text=wrap)}
    if bg:
        cell_style["fill"] = PatternFill("solid", fgColor=bg)
    return cell_style

def apply(ws, cell, styles):
    c = ws[cell] if isinstance(cell, str) else cell
    for attr, val in styles.items():
        setattr(c, attr, val)

def thin_border(top=False, bottom=False, left=False, right=False):
    s = Side(style="thin", color="BFBFBF")
    t = s if top else None; b = s if bottom else None
    l = s if left else None; r = s if right else None
    return Border(top=t, bottom=b, left=l, right=r)

def set_row_style(ws, row, bg_color, cols_range):
    fill = PatternFill("solid", fgColor=bg_color)
    for col in cols_range:
        ws.cell(row=row, column=col).fill = fill

def border_all(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="BFBFBF")
    for r in range(min_row, max_row+1):
        for c in range(min_col, max_col+1):
            ws.cell(row=r, column=c).border = Border(
                top=thin, bottom=thin, left=thin, right=thin)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1: FICHA TÉCNICA
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "FICHA TÉCNICA"
ws1.sheet_view.showGridLines = False

ws1.column_dimensions["A"].width = 14
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 42
ws1.column_dimensions["D"].width = 26
ws1.column_dimensions["E"].width = 22
ws1.column_dimensions["F"].width = 32
ws1.column_dimensions["G"].width = 22
ws1.column_dimensions["H"].width = 22

# ── MAIN TITLE (row 1-2) ───────────────────────────────────────────────────
ws1.merge_cells("A1:H2")
t = ws1["A1"]
t.value = "BASE DE DATOS — TEMA 1: CONECTIVIDAD VIAL Y DESNUTRICIÓN CRÓNICA INFANTIL EN HUÁNUCO"
t.font = Font(name="Arial", bold=True, size=14, color=WHITE)
t.fill = PatternFill("solid", fgColor=DARK_BLUE)
t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws1.row_dimensions[1].height = 28
ws1.row_dimensions[2].height = 16

# ── SUBTITLE / model type ──────────────────────────────────────────────────
ws1.merge_cells("A3:H3")
s = ws1["A3"]
s.value = "Modelo de Regresión Múltiple — Corte Transversal Distrital 2017-2022  (32 distritos de Huánuco)"
s.font = Font(name="Arial", italic=True, size=11, color=DARK_BLUE)
s.fill = PatternFill("solid", fgColor="DCE6F1")
s.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[3].height = 18

# ── EQUATION ──────────────────────────────────────────────────────────────
ws1.merge_cells("A4:H4")
eq = ws1["A4"]
eq.value = "ECUACIÓN DEL MODELO ECONOMÉTRICO"
eq.font = Font(name="Arial", bold=True, size=11, color=WHITE)
eq.fill = PatternFill("solid", fgColor=MID_BLUE)
eq.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[4].height = 20

ws1.merge_cells("A5:H5")
eqv = ws1["A5"]
eqv.value = ("DESNUT_t  =  β₀  +  β₁·VIA_INDEX_t  +  β₂·AGUA_t  "
             "+  β₃·DENS_SALUD_t  +  β₄·ANALF_FEM_t  +  ε_t")
eqv.font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)
eqv.fill = PatternFill("solid", fgColor="EBF3FB")
eqv.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[5].height = 22

# ── VARIABLE TABLE HEADER ──────────────────────────────────────────────────
ws1.merge_cells("A6:H6")
vh = ws1["A6"]
vh.value = "FICHA DE VARIABLES"
vh.font = Font(name="Arial", bold=True, size=11, color=WHITE)
vh.fill = PatternFill("solid", fgColor=DARK_BLUE)
vh.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[6].height = 20

hdrs7 = ["ROL", "CÓDIGO", "DESCRIPCIÓN COMPLETA", "UNIDAD DE MEDIDA",
         "FUENTE", "URL / RUTA DE DESCARGA", "FRECUENCIA ORIGINAL",
         "TRANSFORMACIÓN"]
for col, h in enumerate(hdrs7, 1):
    c = ws1.cell(row=7, column=col, value=h)
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=MID_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
ws1.row_dimensions[7].height = 30

# ── VARIABLE ROWS ──────────────────────────────────────────────────────────
vars_data = [
    ("Y",  "DESNUT",
     "Tasa de desnutrición crónica infantil por distrito (menores de 5 años "
     "con talla-para-edad < -2 DE)",
     "% de niños con DCI sobre total evaluados",
     "MINSA — Sistema SIEN",
     "https://www.minsa.gob.pe → HIS/SIEN → Indicadores Nutricionales por distrito",
     "Anual", "Ninguna (usar directamente)"),
    ("X1", "VIA_INDEX",
     "Índice de accesibilidad vial distrital: km de red vial pavimentada "
     "dividido entre la población distrital × 1,000",
     "km pavim. por cada 1,000 hab.",
     "MTC — Clasificador de Rutas",
     "https://www.gob.pe/mtc → Estadísticas → Clasificador de Rutas / Mapa Vial Huánuco",
     "Anual", "Ratio construido: km_pav / población × 1,000"),
    ("X2", "AGUA_PCT",
     "Porcentaje de hogares con acceso a agua por red pública dentro "
     "de la vivienda o pilón público",
     "% hogares sobre total de hogares del distrito",
     "INEI — Censo Nacional 2017",
     "https://censos2017.inei.gob.pe → Resultados Definitivos → Tabla de Vivienda",
     "Censal (2017)", "Ninguna (dato censal directo)"),
    ("X3", "DENS_SALUD",
     "Densidad de establecimientos de salud por cada 1,000 habitantes "
     "a nivel distrital (EESS 1er y 2do nivel)",
     "Número de EESS por 1,000 hab.",
     "MINSA — RENIPRESS",
     "https://www.minsa.gob.pe → RENIPRESS → Directorio Nacional de EESS → filtrar Huánuco",
     "Anual", "Ratio: N° EESS / población × 1,000"),
    ("X4", "ANALF_FEM",
     "Tasa de analfabetismo femenino distrital: porcentaje de mujeres "
     "de 15 a más años que no saben leer ni escribir",
     "% mujeres analfabetas sobre total mujeres 15+ años",
     "INEI — Censo Nacional 2017",
     "https://censos2017.inei.gob.pe → Resultados Definitivos → Tabla Educación por distrito",
     "Censal (2017)", "Ninguna (dato censal directo)"),
]

row_colors = [LIGHT_BLUE, LIGHT_GREEN, LIGHT_GREEN, LIGHT_GREEN, LIGHT_GREEN]
for i, (rol, cod, desc, unid, fuente, url, freq, transf) in enumerate(vars_data):
    r = 8 + i
    vals = [rol, cod, desc, unid, fuente, url, freq, transf]
    bg = row_colors[i]
    for col, val in enumerate(vals, 1):
        c = ws1.cell(row=r, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", size=10,
                      bold=(col <= 2),
                      color=(BLUE_TEXT if col == 1 else BLACK_TEXT))
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if col <= 2 else "left")
    ws1.row_dimensions[r].height = 48

border_all(ws1, 7, 12, 1, 8)

# ── SIGNS EXPECTED ─────────────────────────────────────────────────────────
ws1.row_dimensions[13].height = 8
ws1.merge_cells("A14:H14")
sh = ws1["A14"]
sh.value = "SIGNOS ESPERADOS DE LOS COEFICIENTES"
sh.font = Font(name="Arial", bold=True, size=11, color=WHITE)
sh.fill = PatternFill("solid", fgColor=DARK_BLUE)
sh.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[14].height = 20

sign_hdrs = ["COEFICIENTE", "SIGNO ESPERADO", "JUSTIFICACIÓN ECONÓMICA"]
for col, h in enumerate(sign_hdrs, 1):
    c = ws1.cell(row=15, column=col)
    c.value = h
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=MID_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[15].height = 20
ws1.merge_cells("C15:H15")
ws1.merge_cells("A15:A15")
ws1.merge_cells("B15:B15")

signs = [
    ("β₁ (VIA_INDEX)", "Negativo (–)",
     "Mayor accesibilidad vial → menor tiempo de traslado a centros de salud y mercados → mejor estado nutricional infantil"),
    ("β₂ (AGUA_PCT)",  "Negativo (–)",
     "Mayor cobertura de agua potable → reducción de enfermedades diarreicas e infecciones → menor desnutrición crónica"),
    ("β₃ (DENS_SALUD)", "Negativo (–)",
     "Mayor densidad de EESS → mayor cobertura de controles CRED y suplementación → menor prevalencia DCI"),
    ("β₄ (ANALF_FEM)", "Positivo (+)",
     "Mayor analfabetismo femenino → menor conocimiento materno sobre nutrición y prácticas de cuidado → mayor desnutrición"),
]
for i, (coef, sgn, just) in enumerate(signs):
    r = 16 + i
    bg = LIGHT_GRAY if i % 2 == 0 else WHITE
    c1 = ws1.cell(row=r, column=1, value=coef)
    c2 = ws1.cell(row=r, column=2, value=sgn)
    ws1.merge_cells(f"C{r}:H{r}")
    c3 = ws1.cell(row=r, column=3, value=just)
    for c in [c1, c2, c3]:
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(name="Arial", size=10,
                      bold=(c == c1),
                      color=("C00000" if "–" in str(sgn) else
                             ("375623" if "+" in str(sgn) else BLACK_TEXT)))
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="center" if c in [c1, c2] else "left")
    ws1.row_dimensions[r].height = 36

border_all(ws1, 15, 19, 1, 8)

# ── METHODOLOGICAL NOTES ──────────────────────────────────────────────────
ws1.row_dimensions[20].height = 8
ws1.merge_cells("A21:H21")
nh = ws1["A21"]
nh.value = "NOTAS METODOLÓGICAS"
nh.font = Font(name="Arial", bold=True, size=11, color=WHITE)
nh.fill = PatternFill("solid", fgColor=DARK_BLUE)
nh.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[21].height = 20

notes = [
    "1. Unidad de análisis: 32 distritos del departamento de Huánuco. Corte transversal para los años 2017 y/o 2022 según disponibilidad de datos censales y SIEN.",
    "2. La variable Y (DESNUT) proviene del SIEN-MINSA. Descargar el consolidado anual por distrito para Huánuco en el portal HIS.",
    "3. VIA_INDEX (X1) se construye dividiendo los kilómetros de red vial pavimentada del distrito (MTC) entre la población proyectada del distrito × 1,000.",
    "4. AGUA_PCT (X2) y ANALF_FEM (X4) provienen del Censo Nacional 2017 (INEI). Disponibles en censos2017.inei.gob.pe con desagregación distrital.",
    "5. DENS_SALUD (X3) se construye dividiendo el número de EESS del distrito (RENIPRESS-MINSA) entre la población proyectada del distrito × 1,000.",
    "6. Verificar normalidad de los residuos (Jarque-Bera) y ausencia de heteroscedasticidad (prueba de White o Breusch-Pagan). Con corte transversal, el riesgo principal es la heteroscedasticidad.",
    "7. Estimar con MCO. Evaluar multicolinealidad con VIF (< 10). Si VIF > 10 entre X2 y X4, considerar eliminar la menos significativa o aplicar componentes principales.",
    "8. Software recomendado: EViews, Stata, R (paquete lm + lmtest + car) o SPSS. Se puede usar Excel con el complemento Analysis ToolPak para regresión básica.",
]
for i, note in enumerate(notes):
    r = 22 + i
    ws1.merge_cells(f"A{r}:H{r}")
    c = ws1["A" + str(r)]
    c.value = note
    bg = YELLOW_HL if i % 2 == 0 else "FFFDE7"
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    ws1.row_dimensions[r].height = 28

border_all(ws1, 22, 29, 1, 8)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2: BASE DE DATOS
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("BASE DE DATOS")
ws2.sheet_view.showGridLines = False

col_widths = {"A": 24, "B": 16, "C": 18, "D": 18, "E": 18, "F": 18,
              "G": 18, "H": 18}
for col, w in col_widths.items():
    ws2.column_dimensions[col].width = w

# title
ws2.merge_cells("A1:H1")
t2 = ws2["A1"]
t2.value = "BASE DE DATOS — TEMA 1: CONECTIVIDAD VIAL Y DESNUTRICIÓN CRÓNICA INFANTIL EN HUÁNUCO"
t2.font = Font(name="Arial", bold=True, size=13, color=WHITE)
t2.fill = PatternFill("solid", fgColor=DARK_BLUE)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 26

ws2.merge_cells("A2:H2")
sub2 = ws2["A2"]
sub2.value = ("Corte transversal 32 distritos de Huánuco | Año de referencia: 2017-2022 | "
              "Fuentes: MINSA-SIEN, MTC, INEI-Censo 2017, RENIPRESS | "
              "Completar celdas en azul con datos descargados")
sub2.font = Font(name="Arial", italic=True, size=10, color=DARK_BLUE)
sub2.fill = PatternFill("solid", fgColor="DCE6F1")
sub2.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws2.row_dimensions[2].height = 28

# ── GROUP HEADERS (row 3) ──────────────────────────────────────────────────
ws2.merge_cells("A3:B3")
g1 = ws2["A3"]
g1.value = "IDENTIFICACIÓN"
g1.font = Font(name="Arial", bold=True, size=10, color=WHITE)
g1.fill = PatternFill("solid", fgColor=MID_BLUE)
g1.alignment = Alignment(horizontal="center", vertical="center")

ws2["C3"].value = "VARIABLE DEPENDIENTE"
ws2["C3"].font = Font(name="Arial", bold=True, size=10, color=WHITE)
ws2["C3"].fill = PatternFill("solid", fgColor="1A5276")
ws2["C3"].alignment = Alignment(horizontal="center", vertical="center")

ws2.merge_cells("D3:G3")
g3 = ws2["D3"]
g3.value = "VARIABLES INDEPENDIENTES"
g3.font = Font(name="Arial", bold=True, size=10, color=WHITE)
g3.fill = PatternFill("solid", fgColor="1D6A96")
g3.alignment = Alignment(horizontal="center", vertical="center")

ws2.merge_cells("H3:H3")
g4 = ws2["H3"]
g4.value = "VARIABLE AUXILIAR"
g4.font = Font(name="Arial", bold=True, size=10, color=WHITE)
g4.fill = PatternFill("solid", fgColor="555555")
g4.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[3].height = 22

# ── COLUMN HEADERS (row 4) ────────────────────────────────────────────────
col_hdrs = [
    ("A4", "DISTRITO", DARK_BLUE),
    ("B4", "PROVINCIA", DARK_BLUE),
    ("C4", "Y  DESNUT\n(% DCI menores 5 años)", "1A5276"),
    ("D4", "X1  VIA_INDEX\n(km pav./1,000 hab.)", "1D6A96"),
    ("E4", "X2  AGUA_PCT\n(% hogares c/agua red)", "1D6A96"),
    ("F4", "X3  DENS_SALUD\n(EESS/1,000 hab.)", "1D6A96"),
    ("G4", "X4  ANALF_FEM\n(% mujeres analfab.)", "1D6A96"),
    ("H4", "LN_VIA\n(log X1)", "555555"),
]
for cell_ref, label, bg in col_hdrs:
    c = ws2[cell_ref]
    c.value = label
    c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center",
                            wrap_text=True)
ws2.row_dimensions[4].height = 40

# ── DATA ROWS (32 distritos de Huánuco) ───────────────────────────────────
distritos = [
    ("Huánuco",            "Huánuco"),
    ("Amarilis",           "Huánuco"),
    ("Pillco Marca",       "Huánuco"),
    ("Chinchao",           "Huánuco"),
    ("Churubamba",         "Huánuco"),
    ("Margos",             "Huánuco"),
    ("Quisqui",            "Huánuco"),
    ("San Francisco de Cayrán", "Huánuco"),
    ("Santa María del Valle",   "Huánuco"),
    ("Yacus",              "Huánuco"),
    ("Ambo",               "Ambo"),
    ("Conchamarca",        "Ambo"),
    ("San Francisco",      "Ambo"),
    ("Tomay Kichwa",       "Ambo"),
    ("Huacaybamba",        "Huacaybamba"),
    ("Canchabamba",        "Huacaybamba"),
    ("Cochabamba",         "Huacaybamba"),
    ("Pinra",              "Huacaybamba"),
    ("Huamalíes",          "Huamalíes"),  # capital: Llata
    ("Arancay",            "Huamalíes"),
    ("Chavinillo",         "Yarowilca"),
    ("Choras",             "Yarowilca"),
    ("Jacas Grande",       "Yarowilca"),
    ("Obas",               "Yarowilca"),
    ("Pampamarca",         "Yarowilca"),
    ("Leoncio Prado",      "Leoncio Prado"),  # capital: Tingo María
    ("José Crespo y Castillo", "Leoncio Prado"),
    ("Luyando",            "Leoncio Prado"),
    ("Mariano Dámaso Beraún", "Leoncio Prado"),
    ("Rupa-Rupa",          "Leoncio Prado"),
    ("Puerto Inca",        "Puerto Inca"),
    ("Honoria",            "Puerto Inca"),
]

for i, (distrito, provincia) in enumerate(distritos):
    r = 5 + i
    bg = LIGHT_BLUE if i % 2 == 0 else WHITE
    ws2.cell(row=r, column=1, value=distrito).fill = PatternFill("solid", fgColor=bg)
    ws2.cell(row=r, column=2, value=provincia).fill = PatternFill("solid", fgColor=bg)
    for col in [1, 2]:
        ws2.cell(row=r, column=col).font = Font(name="Arial", size=10, bold=False)
        ws2.cell(row=r, column=col).alignment = Alignment(vertical="center",
                                                           wrap_text=True)
    # Blue input cells (Y and X variables)
    for col in range(3, 8):
        c = ws2.cell(row=r, column=col)
        c.fill = PatternFill("solid", fgColor="EBF3FB")
        c.font = Font(name="Arial", size=10, color=BLUE_TEXT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.00"
    # LN_VIA formula (gray, auto-calc)
    ln_cell = ws2.cell(row=r, column=8)
    ln_cell.value = f"=IF(D{r}>0,LN(D{r}),\"\")"
    ln_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ln_cell.font = Font(name="Arial", size=10, color=BLACK_TEXT)
    ln_cell.alignment = Alignment(horizontal="center", vertical="center")
    ln_cell.number_format = "0.0000"
    ws2.row_dimensions[r].height = 18

# ── STATISTICS FOOTER ─────────────────────────────────────────────────────
stat_start = 5 + len(distritos)  # row 37
ws2.row_dimensions[stat_start].height = 6

stats = [
    ("N (obs.)",  None, f"=COUNTA(A5:A{stat_start-1})", 
     "=COUNTA(D5:D{})".format(stat_start-1),
     "=COUNTA(E5:E{})".format(stat_start-1),
     "=COUNTA(F5:F{})".format(stat_start-1),
     "=COUNTA(G5:G{})".format(stat_start-1),
     "=COUNTA(H5:H{})".format(stat_start-1)),
    ("Media",     None,
     f"=IFERROR(AVERAGE(C5:C{stat_start-1}),\"\")",
     f"=IFERROR(AVERAGE(D5:D{stat_start-1}),\"\")",
     f"=IFERROR(AVERAGE(E5:E{stat_start-1}),\"\")",
     f"=IFERROR(AVERAGE(F5:F{stat_start-1}),\"\")",
     f"=IFERROR(AVERAGE(G5:G{stat_start-1}),\"\")",
     f"=IFERROR(AVERAGE(H5:H{stat_start-1}),\"\")"),
    ("Desv. Estándar", None,
     f"=IFERROR(STDEV(C5:C{stat_start-1}),\"\")",
     f"=IFERROR(STDEV(D5:D{stat_start-1}),\"\")",
     f"=IFERROR(STDEV(E5:E{stat_start-1}),\"\")",
     f"=IFERROR(STDEV(F5:F{stat_start-1}),\"\")",
     f"=IFERROR(STDEV(G5:G{stat_start-1}),\"\")",
     f"=IFERROR(STDEV(H5:H{stat_start-1}),\"\")"),
    ("Mín.", None,
     f"=IFERROR(MIN(C5:C{stat_start-1}),0)",
     f"=IFERROR(MIN(D5:D{stat_start-1}),0)",
     f"=IFERROR(MIN(E5:E{stat_start-1}),0)",
     f"=IFERROR(MIN(F5:F{stat_start-1}),0)",
     f"=IFERROR(MIN(G5:G{stat_start-1}),0)",
     f"=IFERROR(MIN(H5:H{stat_start-1}),0)"),
    ("Máx.", None,
     f"=IFERROR(MAX(C5:C{stat_start-1}),0)",
     f"=IFERROR(MAX(D5:D{stat_start-1}),0)",
     f"=IFERROR(MAX(E5:E{stat_start-1}),0)",
     f"=IFERROR(MAX(F5:F{stat_start-1}),0)",
     f"=IFERROR(MAX(G5:G{stat_start-1}),0)",
     f"=IFERROR(MAX(H5:H{stat_start-1}),0)"),
]

for j, row_vals in enumerate(stats):
    r = stat_start + 1 + j
    label = row_vals[0]
    ws2.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=10)
    ws2.cell(row=r, column=1).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws2.cell(row=r, column=2, value="").fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    for col, val in enumerate(row_vals[2:], 3):
        c = ws2.cell(row=r, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        c.font = Font(name="Arial", size=10, bold=False)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.00"
    ws2.row_dimensions[r].height = 18

# Legend
legend_row = stat_start + 7
ws2.merge_cells(f"A{legend_row}:H{legend_row}")
leg = ws2[f"A{legend_row}"]
leg.value = ("LEYENDA DE COLORES:  🟢 Verde-azul = Variable dependiente (Y) — ingresar datos manualmente   "
             "🔵 Azul = Variables independientes (X) — ingresar datos manualmente   "
             "⬜ Gris = Fórmulas calculadas automáticamente")
leg.font = Font(name="Arial", size=9, italic=True)
leg.fill = PatternFill("solid", fgColor=YELLOW_HL)
leg.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws2.row_dimensions[legend_row].height = 28

border_all(ws2, 3, stat_start + 5, 1, 8)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3: GUÍA DE DESCARGA
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("GUÍA DE DESCARGA")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width = 8
ws3.column_dimensions["B"].width = 30
ws3.column_dimensions["C"].width = 72

ws3.merge_cells("A1:C1")
t3 = ws3["A1"]
t3.value = "GUÍA PASO A PASO PARA DESCARGA DE DATOS — TEMA 1"
t3.font = Font(name="Arial", bold=True, size=13, color=WHITE)
t3.fill = PatternFill("solid", fgColor=DARK_BLUE)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 26

sections = [
    {
        "title": "Y — DESNUT | Fuente: MINSA — Sistema SIEN (Estado Nutricional)",
        "steps": [
            ("1", "Ir al portal del MINSA",
             "https://www.minsa.gob.pe"),
            ("2", "Ruta de navegación",
             "Inicio → Información de Salud → HIS → SIEN (Sistema de Información del Estado Nutricional)"),
            ("3", "Seleccionar reporte distrital",
             "En el SIEN, busca 'Indicadores Nutricionales por distrito' → Huánuco → Año 2022 (o el más reciente disponible)."),
            ("4", "Dato a extraer",
             "Descarga la tabla de Desnutrición Crónica Infantil (DCI) en menores de 5 años por distrito de Huánuco (%). Columna: '% DCI talla/edad < -2 DE (patrón OMS)'."),
            ("5", "Alternativa — INEI ENDES",
             "Si el SIEN no tiene desagregación distrital completa, usa la Encuesta Demográfica y de Salud Familiar (ENDES) del INEI: proyectos.inei.gob.pe/microdatos → ENDES → filtrar departamento Huánuco."),
        ]
    },
    {
        "title": "X1 — VIA_INDEX | Fuente: MTC — Clasificador de Rutas y Mapa Vial",
        "steps": [
            ("1", "Ir al portal del MTC",
             "https://www.gob.pe/mtc → sección Estadísticas → Transporte Terrestre → Infraestructura Vial"),
            ("2", "Descargar Clasificador de Rutas",
             "Busca el archivo Excel 'Clasificador de Rutas del SINAC'. Filtra por departamento Huánuco y anota km por tipo de superficie (pavimentado / afirmado / sin afirmar) a nivel distrital."),
            ("3", "Mapa Vial Departamental",
             "Complementar con el 'Mapa Vial del Departamento de Huánuco' (PDF/SHP). Disponible en MTC → Portal Geo → capas de red vial."),
            ("4", "Construir el índice",
             "VIA_INDEX = km_vial_pavimentada_distrito / Población_distrital × 1,000. La población distrital 2017 o 2022 se obtiene de INEI → Estimaciones y Proyecciones."),
            ("5", "Fuente de población",
             "INEI: https://www.inei.gob.pe → Estadísticas → Población y Vivienda → Estimaciones y Proyecciones de Población → tabla por departamento/provincia/distrito."),
        ]
    },
    {
        "title": "X2 — AGUA_PCT | Fuente: INEI — Censo Nacional de Población y Vivienda 2017",
        "steps": [
            ("1", "Ir a Censos 2017",
             "https://censos2017.inei.gob.pe → sección 'Resultados Definitivos'"),
            ("2", "Seleccionar tomo de vivienda",
             "Descargar el Tomo de Vivienda para Huánuco (archivo PDF o Excel con tablas estadísticas). Busca la tabla: 'Viviendas particulares con ocupantes presentes, por tipo de abastecimiento de agua'."),
            ("3", "Dato a extraer por distrito",
             "Para cada distrito de Huánuco, extraer: % hogares con agua por 'red pública dentro de la vivienda' + 'red pública fuera de la vivienda' + 'pilón o pileta pública'. Sumar los tres valores = AGUA_PCT."),
            ("4", "Alternativa — ENAHO",
             "INEI ENAHO (módulo de vivienda): proyectos.inei.gob.pe/microdatos → ENAHO → Módulo 100 (Características de la vivienda) → variable P1121 (abastecimiento de agua) → filtrar Huánuco."),
        ]
    },
    {
        "title": "X3 — DENS_SALUD | Fuente: MINSA — RENIPRESS (Directorio de EESS)",
        "steps": [
            ("1", "Ir al RENIPRESS",
             "https://www.minsa.gob.pe → Datos Abiertos → Registro Nacional de EESS (RENIPRESS) → Descargar base de datos completa en Excel."),
            ("2", "Filtrar por Huánuco",
             "Abrir el archivo Excel del RENIPRESS. Filtrar columna 'Departamento' = HUANUCO. Luego filtrar por 'Categoría' solo niveles I y II (primer y segundo nivel de atención)."),
            ("3", "Contar EESS por distrito",
             "Con tabla dinámica o CONTARSI, contar el número de establecimientos activos por distrito. Resultado: N° EESS por cada distrito."),
            ("4", "Construir la densidad",
             "DENS_SALUD = N°_EESS_distrito / Población_distrital × 1,000. Usar la misma población que en X1 (INEI proyecciones)."),
        ]
    },
    {
        "title": "X4 — ANALF_FEM | Fuente: INEI — Censo Nacional 2017",
        "steps": [
            ("1", "Ir a Censos 2017",
             "https://censos2017.inei.gob.pe → Resultados Definitivos → Tomo de Educación para Huánuco."),
            ("2", "Dato a extraer",
             "Busca la tabla: 'Población de 15 y más años de edad por sexo, según nivel educativo y condición de alfabetismo'. Extraer para cada distrito: mujeres analfabetas / total mujeres 15+ años × 100 = ANALF_FEM."),
            ("3", "Alternativa — INEI datos abiertos",
             "datosabiertos.gob.pe → INEI → Indicadores de educación por distrito → buscar tasa de analfabetismo femenino departamento Huánuco."),
            ("4", "Validación",
             "La tasa departamental de Huánuco (referencia): aproximadamente 14%-18% de analfabetismo femenino según Censo 2017. Distritos rurales de sierra pueden superar el 30%. Verificar coherencia de los valores extraídos."),
        ]
    },
]

current_row = 2
for sec in sections:
    # Section header
    ws3.merge_cells(f"A{current_row}:C{current_row}")
    sh = ws3[f"A{current_row}"]
    sh.value = sec["title"]
    sh.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    sh.fill = PatternFill("solid", fgColor=MID_BLUE)
    sh.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws3.row_dimensions[current_row].height = 24
    current_row += 1

    # Column sub-headers
    for col, label in [(1, "PASO"), (2, "ACCIÓN"), (3, "DETALLE")]:
        c = ws3.cell(row=current_row, column=col, value=label)
        c.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor="3D85C8")
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[current_row].height = 18
    current_row += 1

    for j, (paso, accion, detalle) in enumerate(sec["steps"]):
        bg = LIGHT_GRAY if j % 2 == 0 else WHITE
        c1 = ws3.cell(row=current_row, column=1, value=paso)
        c2 = ws3.cell(row=current_row, column=2, value=accion)
        c3 = ws3.cell(row=current_row, column=3, value=detalle)
        for c, aln in [(c1, "center"), (c2, "left"), (c3, "left")]:
            c.fill = PatternFill("solid", fgColor=bg)
            c.font = Font(name="Arial", size=10,
                          bold=(c == c2))
            c.alignment = Alignment(horizontal=aln, vertical="center",
                                    wrap_text=True, indent=(1 if c != c1 else 0))
        ws3.row_dimensions[current_row].height = 44
        current_row += 1

    # small gap
    ws3.row_dimensions[current_row].height = 6
    current_row += 1

border_all(ws3, 2, current_row - 2, 1, 3)


# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4: MODELO ECONOMÉTRICO
# ═══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("MODELO ECONOMÉTRICO")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 30
ws4.column_dimensions["B"].width = 72

ws4.merge_cells("A1:B1")
t4 = ws4["A1"]
t4.value = "ESQUEMA DEL MODELO ECONOMÉTRICO — TEMA 1"
t4.font = Font(name="Arial", bold=True, size=13, color=WHITE)
t4.fill = PatternFill("solid", fgColor=DARK_BLUE)
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 26

def section_block(ws, title, rows_data, start_row, hdr_bg=MID_BLUE):
    ws.merge_cells(f"A{start_row}:B{start_row}")
    h = ws[f"A{start_row}"]
    h.value = title
    h.font = Font(name="Arial", bold=True, size=11, color=WHITE)
    h.fill = PatternFill("solid", fgColor=hdr_bg)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[start_row].height = 22
    start_row += 1
    for j, (label, val) in enumerate(rows_data):
        bg = LIGHT_GRAY if j % 2 == 0 else WHITE
        c1 = ws.cell(row=start_row, column=1, value=label)
        c1.fill = PatternFill("solid", fgColor=bg)
        c1.font = Font(name="Arial", bold=True, size=10)
        c1.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        c2 = ws.cell(row=start_row, column=2, value=val)
        c2.fill = PatternFill("solid", fgColor=bg)
        c2.font = Font(name="Arial", size=10)
        c2.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[start_row].height = 36
        start_row += 1
    border_all(ws, start_row - len(rows_data) - 1, start_row - 1, 1, 2)
    return start_row + 1

r = 2
r = section_block(ws4, "ESPECIFICACIÓN DEL MODELO", [
    ("Tipo de modelo",
     "Regresión lineal múltiple por Mínimos Cuadrados Ordinarios (MCO) — Corte Transversal"),
    ("Forma funcional",
     "Log-lineal parcial: Y en niveles; X1 en logaritmo natural (LN) para reducir asimetría del índice vial; X2, X3, X4 en niveles porcentuales"),
    ("Ecuación estimable",
     "DESNUT_i = β₀ + β₁·LN(VIA_INDEX_i) + β₂·AGUA_PCT_i + β₃·DENS_SALUD_i + β₄·ANALF_FEM_i + ε_i"),
    ("Unidad de análisis",
     "Distritos del departamento de Huánuco (N = 32)"),
    ("Número de parámetros",
     "5 (β₀, β₁, β₂, β₃, β₄)"),
    ("Grados de libertad",
     "27 (adecuado para MCO con 4 regresores y 32 observaciones)"),
], r)

r = section_block(ws4, "PASOS DE ESTIMACIÓN", [
    ("Paso 1 — Análisis descriptivo",
     "Calcular media, desviación estándar, mínimo y máximo de cada variable. Construir histogramas de Y (DESNUT) para evaluar distribución. Mapear la variable dependiente por distrito para detectar clusters espaciales."),
    ("Paso 2 — Análisis de correlaciones",
     "Construir matriz de correlaciones entre X1 (LN), X2, X3, X4. Si |r| > 0,90 entre dos regresores → posible multicolinealidad severa. Calcular VIF (Variance Inflation Factor): VIF < 10 es aceptable."),
    ("Paso 3 — Estimación MCO",
     "Estimar la ecuación por OLS. Obtener: coeficientes β̂, errores estándar, t-estadísticos, p-valores, R², R² ajustado y F-estadístico global del modelo."),
    ("Paso 4 — Diagnóstico de residuos (corte transversal)",
     "a) Heteroscedasticidad: prueba de White (H₀: residuos homosc.). En datos de corte transversal distrital, la heteroscedasticidad es el problema más frecuente. Si se rechaza H₀ → usar errores estándar robustos (White/HC).\nb) Normalidad: prueba Jarque-Bera sobre residuos.\nc) Detección de observaciones influyentes: Distancia de Cook, apalancamiento (leverage). Con N=32 es crucial identificar outliers."),
    ("Paso 5 — Interpretación de coeficientes",
     "β₁: disminución en pp de DESNUT ante un aumento del 1% en VIA_INDEX (elasticidad parcial).\nβ₂: cambio en pp de DESNUT ante un aumento de 1 pp en AGUA_PCT.\nβ₃: cambio en pp de DESNUT ante el aumento de 1 EESS por mil habitantes.\nβ₄: cambio en pp de DESNUT ante un aumento de 1 pp en ANALF_FEM."),
    ("Paso 6 — Modelo ampliado (opcional)",
     "Incluir término de interacción β₅·(VIA_INDEX × ANALF_FEM) para testear si el efecto de la accesibilidad vial es heterogéneo según el nivel de capital humano femenino distrital."),
], r)

r = section_block(ws4, "CRITERIOS DE BONDAD DE AJUSTE", [
    ("R² ajustado",
     "Debe ser > 0,50 para considerarse aceptable con N=32 en datos de corte transversal."),
    ("F-estadístico (global)",
     "p-valor < 0,05 → el modelo en su conjunto es significativo."),
    ("t-estadísticos",
     "p-valor < 0,05 por variable → coeficiente individualmente significativo."),
    ("Prueba de White",
     "p-valor > 0,05 indica homocedasticidad. Si < 0,05, usar errores estándar robustos (HC3)."),
    ("VIF",
     "VIF < 10 para cada regresor indica ausencia de multicolinealidad severa. VIF < 5 es ideal."),
    ("AIC / BIC",
     "Útiles para comparar especificaciones alternativas del modelo (con/sin interacciones, con/sin transformaciones logarítmicas)."),
], r)

r = section_block(ws4, "HIPÓTESIS DE LA INVESTIGACIÓN", [
    ("Hipótesis general",
     "La accesibilidad vial, la cobertura de agua potable, la densidad de establecimientos de salud y el capital humano femenino determinan significativamente la tasa de desnutrición crónica infantil entre los distritos del departamento de Huánuco."),
    ("H₁ (X1: VIA_INDEX)",
     "A mayor índice de accesibilidad vial distrital, menor tasa de desnutrición crónica infantil, al mejorar el acceso a centros de salud y mercados de alimentos (β₁ < 0)."),
    ("H₂ (X2: AGUA_PCT)",
     "A mayor porcentaje de hogares con acceso a agua potable, menor tasa de desnutrición, por reducción de enfermedades gastrointestinales que deterioran la absorción de nutrientes (β₂ < 0)."),
    ("H₃ (X3: DENS_SALUD)",
     "A mayor densidad de establecimientos de salud, menor tasa de desnutrición, al ampliar la cobertura de controles de Crecimiento y Desarrollo (CRED) y programas de suplementación con micronutrientes (β₃ < 0)."),
    ("H₄ (X4: ANALF_FEM)",
     "A mayor tasa de analfabetismo femenino distrital, mayor tasa de desnutrición infantil, dada la correlación entre la educación materna y las prácticas de alimentación, higiene y cuidado (β₄ > 0)."),
], r)

# ── Final freeze panes on data sheet ──────────────────────────────────────
ws2.freeze_panes = "A5"

# ── SAVE ──────────────────────────────────────────────────────────────────
out = "Tema1_DesnutricionVial_Huanuco.xlsx"
wb.save(out)
print("Saved:", out)